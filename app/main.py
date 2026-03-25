import asyncio
import calendar
import json
import os
import re
import smtplib
import ssl
import time
from email.message import EmailMessage
try:
    import winsound
except Exception:
    winsound = None
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Optional, Dict, List, Tuple

from fastapi import FastAPI, HTTPException, Request, Form
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from .db import (
    init_db,
    upsert_person,
    get_person_by_uid,
    list_people,
    list_people_filtered,
    list_people_all,
    add_attendance,
    list_attendance,
    list_attendance_filtered,
    list_attendance_all,
)
from .reader import CardReaderService, CardRead

app = FastAPI(title="Lector Omnikey", version="0.1.0")
app.add_middleware(SessionMiddleware, secret_key=os.environ.get("SESSION_SECRET", "change-me"))

static_dir = Path(__file__).parent / "static"
app.mount("/static", StaticFiles(directory=static_dir), name="static")

_project_root = Path(__file__).resolve().parents[1]

_reader_service: Optional[CardReaderService] = None
_loop: Optional[asyncio.AbstractEventLoop] = None
_last_read: Optional[Dict[str, str]] = None
_reader_state: Dict[str, Optional[str]] = {"present": False, "last_seen_at": None}
_control_last_ping_monotonic: float = 0.0
_control_ping_ttl_seconds: float = 4.0
_report_task: Optional[asyncio.Task] = None
_report_stop_event: Optional[asyncio.Event] = None
_report_config_path = Path(os.environ.get("REPORT_CONFIG_PATH", str(_project_root / "data" / "report_config.json")))
_report_state_path = Path(os.environ.get("REPORT_STATE_PATH", str(_project_root / "data" / "report_state.json")))


class PersonIn(BaseModel):
    uid: str = Field(..., min_length=4)
    first_name: str = Field(..., min_length=1)
    last_name: str = Field(..., min_length=1)
    id_number: str = Field(..., min_length=3)
    phone: Optional[str] = None
    area: Optional[str] = None


def _is_admin(request: Request) -> bool:
    return bool(request.session.get("admin"))


def _require_admin(request: Request) -> None:
    if not _is_admin(request):
        raise HTTPException(status_code=401, detail="Admin login required")


def _autosize_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def _format_dt(value: str) -> str:
    try:
        dt = datetime.fromisoformat(value)
        return dt.astimezone().strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return value


def _normalize_query(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    value = value.strip()
    return value or None


def _parse_dt_param(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(value)
        if dt.tzinfo is None:
            local_tz = datetime.now().astimezone().tzinfo
            dt = dt.replace(tzinfo=local_tz)
        return dt.astimezone(timezone.utc).isoformat()
    except Exception:
        return None


def _parse_month_param(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    value = value.strip()
    if re.fullmatch(r"\d{4}-\d{2}", value):
        return value
    return None


def _load_report_config() -> Dict[str, object]:
    defaults: Dict[str, object] = {
        "enabled": os.environ.get("REPORT_EMAIL_ENABLED", "0") == "1",
        "recipient_email": os.environ.get("REPORT_RECIPIENT_EMAIL", "").strip(),
        "sender_email": os.environ.get("REPORT_SENDER_EMAIL", "").strip(),
        "sender_password": os.environ.get("REPORT_SENDER_PASSWORD", "").strip(),
        "smtp_host": os.environ.get("REPORT_SMTP_HOST", "smtp.gmail.com").strip(),
        "smtp_port": int(os.environ.get("REPORT_SMTP_PORT", "587")),
        "use_tls": os.environ.get("REPORT_SMTP_USE_TLS", "1") == "1",
        "send_every_days": int(os.environ.get("REPORT_EVERY_DAYS", "3")),
        "check_interval_minutes": int(os.environ.get("REPORT_CHECK_INTERVAL_MINUTES", "15")),
    }
    if _report_config_path.exists():
        try:
            loaded = json.loads(_report_config_path.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                defaults.update(loaded)
        except Exception as exc:
            print(f"[report] No se pudo leer configuracion {_report_config_path}: {exc}")
    defaults["enabled"] = bool(defaults.get("enabled"))
    defaults["send_every_days"] = max(1, int(defaults.get("send_every_days") or 3))
    defaults["check_interval_minutes"] = max(1, int(defaults.get("check_interval_minutes") or 15))
    defaults["smtp_port"] = int(defaults.get("smtp_port") or 587)
    defaults["use_tls"] = bool(defaults.get("use_tls"))
    return defaults


def _load_report_state() -> Dict[str, object]:
    default_state: Dict[str, object] = {
        "last_control_report_date": None,
        "sent_quincena_keys": [],
    }
    if not _report_state_path.exists():
        return default_state
    try:
        state = json.loads(_report_state_path.read_text(encoding="utf-8"))
        if not isinstance(state, dict):
            return default_state
        # Backward compatibility with previous key name.
        if "last_control_report_date" not in state and "last_people_report_date" in state:
            state["last_control_report_date"] = state.get("last_people_report_date")
        state.setdefault("last_control_report_date", None)
        state.setdefault("sent_quincena_keys", [])
        if not isinstance(state["sent_quincena_keys"], list):
            state["sent_quincena_keys"] = []
        return state
    except Exception as exc:
        print(f"[report] No se pudo leer estado {_report_state_path}: {exc}")
        return default_state


def _save_report_state(state: Dict[str, object]) -> None:
    _report_state_path.parent.mkdir(parents=True, exist_ok=True)
    _report_state_path.write_text(json.dumps(state, indent=2), encoding="utf-8")


def _date_from_iso(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value).date()
    except Exception:
        return None


def _build_people_workbook(rows: List[Dict[str, str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Personas"
    headers = ["UID", "Nombre", "Apellido", "Cedula", "Telefono", "Area", "Creado"]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row["uid"],
                row["first_name"],
                row["last_name"],
                row["id_number"],
                row["phone"],
                row["area"],
                _format_dt(row["created_at"]),
            ]
        )
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        table = Table(displayName="PersonasAuto", ref=f"A1:G{ws.max_row}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)
    _autosize_columns(ws)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _build_attendance_workbook(rows: List[Dict[str, str]], sheet_name: str = "Control") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Control"
    headers = ["UID", "Nombre", "Apellido", "Cedula", "Telefono", "Area", "ATR", "FechaHora"]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row["uid"],
                row["first_name"],
                row["last_name"],
                row["id_number"],
                row["phone"],
                row["area"],
                row["atr"],
                _format_dt(row["read_at"]),
            ]
        )
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        table = Table(displayName="ControlAuto", ref=f"A1:H{ws.max_row}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)
    _autosize_columns(ws)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _send_email_with_attachment(
    cfg: Dict[str, object],
    *,
    subject: str,
    body: str,
    filename: str,
    attachment_bytes: bytes,
) -> None:
    sender = str(cfg.get("sender_email") or "").strip()
    recipient = str(cfg.get("recipient_email") or "").strip()
    password = str(cfg.get("sender_password") or "")
    smtp_host = str(cfg.get("smtp_host") or "smtp.gmail.com")
    smtp_port = int(cfg.get("smtp_port") or 587)
    use_tls = bool(cfg.get("use_tls"))

    if not sender or not recipient or not password:
        raise RuntimeError("Falta configuracion de correo (sender, recipient o password)")

    msg = EmailMessage()
    msg["From"] = sender
    msg["To"] = recipient
    msg["Subject"] = subject
    msg.set_content(body)
    msg.add_attachment(
        attachment_bytes,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )

    context = ssl.create_default_context()
    with smtplib.SMTP(smtp_host, smtp_port, timeout=40) as smtp:
        smtp.ehlo()
        if use_tls:
            smtp.starttls(context=context)
            smtp.ehlo()
        smtp.login(sender, password)
        smtp.send_message(msg)


def _quincena_to_send(today: date) -> Optional[Tuple[str, date, date, str]]:
    if today.day == 16:
        start = date(today.year, today.month, 1)
        end = date(today.year, today.month, 15)
        key = f"{today.year:04d}-{today.month:02d}-Q1"
        label = f"01 al 15/{today.month:02d}/{today.year:04d}"
        return key, start, end, label
    if today.day == 1:
        prev_year = today.year
        prev_month = today.month - 1
        if prev_month == 0:
            prev_month = 12
            prev_year -= 1
        last_day = calendar.monthrange(prev_year, prev_month)[1]
        start = date(prev_year, prev_month, 16)
        end = date(prev_year, prev_month, last_day)
        key = f"{prev_year:04d}-{prev_month:02d}-Q2"
        label = f"16 al {last_day:02d}/{prev_month:02d}/{prev_year:04d}"
        return key, start, end, label
    return None


async def _run_report_scheduler(stop_event: asyncio.Event) -> None:
    while not stop_event.is_set():
        try:
            cfg = _load_report_config()
            if cfg.get("enabled"):
                today = datetime.now().date()
                state = _load_report_state()

                last_control_date = _date_from_iso(state.get("last_control_report_date"))
                send_every_days = int(cfg.get("send_every_days") or 3)
                if last_control_date is None or (today - last_control_date).days >= send_every_days:
                    if last_control_date is None:
                        range_start = today - timedelta(days=(send_every_days - 1))
                    else:
                        range_start = last_control_date + timedelta(days=1)
                    range_end = today
                    from_raw = f"{range_start.isoformat()}T00:00:00"
                    to_raw = f"{range_end.isoformat()}T23:59:59"
                    control_rows = await list_attendance_filtered(
                        from_dt=_parse_dt_param(from_raw),
                        to_dt=_parse_dt_param(to_raw),
                        name=None,
                        id_number=None,
                        area=None,
                        uid=None,
                        month_key=None,
                        limit=50000,
                    )
                    control_rows = sorted(control_rows, key=lambda row: str(row.get("read_at") or ""))
                    control_xlsx = _build_attendance_workbook(control_rows, sheet_name="Control 3 dias")
                    subject = (
                        "Control asistencia cada 3 dias - "
                        f"{range_start.isoformat()} a {range_end.isoformat()}"
                    )
                    body = (
                        "Reporte automatico de control de asistencias ordenado por fecha. "
                        f"Periodo: {range_start.isoformat()} al {range_end.isoformat()}. "
                        f"Total: {len(control_rows)} lecturas."
                    )
                    await asyncio.to_thread(
                        _send_email_with_attachment,
                        cfg,
                        subject=subject,
                        body=body,
                        filename=(
                            f"control-3-dias-{range_start.isoformat()}-a-{range_end.isoformat()}.xlsx"
                        ),
                        attachment_bytes=control_xlsx,
                    )
                    state["last_control_report_date"] = today.isoformat()
                    print(
                        "[report] Reporte de control 3 dias enviado "
                        f"({range_start.isoformat()} a {range_end.isoformat()}, {len(control_rows)} lecturas)"
                    )

                quincena = _quincena_to_send(today)
                if quincena is not None:
                    key, start, end, label = quincena
                    sent_keys = [str(k) for k in state.get("sent_quincena_keys", [])]
                    if key not in sent_keys:
                        from_raw = f"{start.isoformat()}T00:00:00"
                        to_raw = f"{end.isoformat()}T23:59:59"
                        attendance_rows = await list_attendance_filtered(
                            from_dt=_parse_dt_param(from_raw),
                            to_dt=_parse_dt_param(to_raw),
                            name=None,
                            id_number=None,
                            area=None,
                            uid=None,
                            month_key=None,
                            limit=50000,
                        )
                        attendance_xlsx = _build_attendance_workbook(
                            attendance_rows,
                            sheet_name="Quincena",
                        )
                        subject = f"Reporte quincenal - {label}"
                        body = (
                            "Reporte automatico de asistencia por quincena. "
                            f"Periodo: {label}. Total: {len(attendance_rows)} lecturas."
                        )
                        await asyncio.to_thread(
                            _send_email_with_attachment,
                            cfg,
                            subject=subject,
                            body=body,
                            filename=f"quincena-{start.isoformat()}-a-{end.isoformat()}.xlsx",
                            attachment_bytes=attendance_xlsx,
                        )
                        sent_keys.append(key)
                        state["sent_quincena_keys"] = sent_keys[-24:]
                        print(
                            "[report] Reporte quincenal enviado "
                            f"({label}, {len(attendance_rows)} lecturas)"
                        )

                _save_report_state(state)
        except Exception as exc:
            print(f"[report] Error en scheduler de reportes: {exc}")

        wait_seconds = int(_load_report_config().get("check_interval_minutes", 15)) * 60
        try:
            await asyncio.wait_for(stop_event.wait(), timeout=wait_seconds)
        except asyncio.TimeoutError:
            continue


@app.on_event("startup")
async def on_startup() -> None:
    global _reader_service, _loop, _report_task, _report_stop_event
    await init_db()
    _loop = asyncio.get_running_loop()

    def handle_card(read: CardRead) -> None:
        global _last_read
        if _loop is None:
            return
        read_at = datetime.now(timezone.utc).isoformat()
        _last_read = {
            "uid": read.uid,
            "atr": read.atr,
            "read_at": read_at,
            "attendance_skipped": None,
            "attendance_message": "Procesando lectura...",
            "allowed_at": None,
        }
        _reader_state["present"] = True
        _reader_state["last_seen_at"] = read_at
        if winsound is not None:
            try:
                winsound.MessageBeep(winsound.MB_OK)
            except Exception:
                pass
        print(f"UID leido: {read.uid} | ATR: {read.atr}")
        control_recently_active = (time.monotonic() - _control_last_ping_monotonic) <= _control_ping_ttl_seconds
        if control_recently_active:
            future = asyncio.run_coroutine_threadsafe(add_attendance(read.uid, read.atr), _loop)

            def _attendance_done(fut) -> None:
                try:
                    result = fut.result()
                    if _last_read is None or _last_read.get("uid") != read.uid:
                        return
                    if isinstance(result, dict) and result.get("skipped"):
                        _last_read["attendance_skipped"] = True
                        _last_read["allowed_at"] = result.get("allowed_at")
                        _last_read["attendance_message"] = "Advertencia: solamente puedes poner una vez la tarjeta."
                        print(
                            "[control] Lectura omitida por ventana de 3 horas "
                            f"para UID {read.uid}. Permitido desde: {result.get('allowed_at')}"
                        )
                    elif result is None:
                        _last_read["attendance_skipped"] = None
                        _last_read["allowed_at"] = None
                        _last_read["attendance_message"] = "Tarjeta no registrada."
                    else:
                        _last_read["attendance_skipped"] = False
                        _last_read["allowed_at"] = None
                        _last_read["attendance_message"] = "Registro realizado correctamente."
                except Exception as exc:
                    print(f"[control] Error registrando asistencia: {exc}")

            future.add_done_callback(_attendance_done)

    def handle_remove() -> None:
        _reader_state["present"] = False

    _reader_service = CardReaderService(on_card=handle_card, on_remove=handle_remove)
    _reader_service.start()

    _report_stop_event = asyncio.Event()
    _report_task = asyncio.create_task(_run_report_scheduler(_report_stop_event))


@app.on_event("shutdown")
async def on_shutdown() -> None:
    global _report_task, _report_stop_event
    if _reader_service:
        _reader_service.stop()
    if _report_stop_event:
        _report_stop_event.set()
    if _report_task:
        try:
            await _report_task
        except Exception:
            pass
        _report_task = None
        _report_stop_event = None


@app.get("/", response_class=HTMLResponse)
async def index() -> HTMLResponse:
    html_path = static_dir / "index.html"
    return HTMLResponse(html_path.read_text(encoding="utf-8"))


@app.get("/admin/login", response_class=HTMLResponse)
async def admin_login_page() -> HTMLResponse:
    html_path = static_dir / "login.html"
    return HTMLResponse(html_path.read_text(encoding="utf-8"))


@app.post("/admin/login")
async def admin_login(request: Request, password: str = Form(...)) -> dict:
    admin_password = os.environ.get("ADMIN_PASSWORD", "12345678")
    if password != admin_password:
        raise HTTPException(status_code=401, detail="Clave incorrecta")
    request.session["admin"] = True
    return {"ok": True, "redirect": "/admin"}


@app.post("/admin/logout")
async def admin_logout(request: Request) -> RedirectResponse:
    request.session.clear()
    return RedirectResponse(url="/", status_code=303)


@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request) -> HTMLResponse:
    _require_admin(request)
    html_path = static_dir / "admin.html"
    return HTMLResponse(html_path.read_text(encoding="utf-8"))


@app.get("/control", response_class=HTMLResponse)
async def control_page() -> HTMLResponse:
    html_path = static_dir / "control.html"
    return HTMLResponse(html_path.read_text(encoding="utf-8"))


@app.post("/api/control/ping")
async def control_ping() -> dict:
    global _control_last_ping_monotonic
    _control_last_ping_monotonic = time.monotonic()
    return {"ok": True}


@app.get("/api/health")
async def health() -> dict:
    status = "ok"
    details = {}
    if _reader_service and _reader_service.init_error:
        status = "error"
        details["reader_error"] = _reader_service.init_error
    return {"status": status, "details": details}


@app.get("/api/latest")
async def latest() -> dict:
    if _last_read is None:
        return {"latest": None}
    person = None
    if _last_read.get("attendance_skipped") is False:
        person = await get_person_by_uid(_last_read["uid"])
    latest_payload = {**_last_read, "person": person, **_reader_state}
    return {"latest": latest_payload}


@app.post("/api/people")
async def create_person(payload: PersonIn, request: Request) -> dict:
    _require_admin(request)
    person = await upsert_person(
        uid=payload.uid,
        first_name=payload.first_name,
        last_name=payload.last_name,
        id_number=payload.id_number,
        phone=payload.phone,
        area=payload.area,
    )
    return {"person": person}


@app.get("/api/people")
async def people(request: Request, limit: int = 100) -> dict:
    _require_admin(request)
    data = await list_people(limit=limit)
    return {"people": data}


@app.get("/api/people/search")
async def people_search(
    request: Request,
    from_dt: Optional[str] = None,
    to_dt: Optional[str] = None,
    name: Optional[str] = None,
    id_number: Optional[str] = None,
    area: Optional[str] = None,
    limit: int = 200,
) -> dict:
    _require_admin(request)
    data = await list_people_filtered(
        from_dt=_parse_dt_param(from_dt),
        to_dt=_parse_dt_param(to_dt),
        name=_normalize_query(name),
        id_number=_normalize_query(id_number),
        area=_normalize_query(area),
        limit=limit,
    )
    return {"people": data}


@app.get("/api/people/{uid}")
async def person_by_uid(uid: str, request: Request) -> dict:
    _require_admin(request)
    person = await get_person_by_uid(uid)
    if person is None:
        raise HTTPException(status_code=404, detail="Person not found")
    return {"person": person}


@app.get("/api/attendance")
async def attendance(
    request: Request,
    from_dt: Optional[str] = None,
    to_dt: Optional[str] = None,
    name: Optional[str] = None,
    id_number: Optional[str] = None,
    area: Optional[str] = None,
    uid: Optional[str] = None,
    month: Optional[str] = None,
    limit: int = 200,
) -> dict:
    _require_admin(request)
    data = await list_attendance_filtered(
        from_dt=_parse_dt_param(from_dt),
        to_dt=_parse_dt_param(to_dt),
        name=_normalize_query(name),
        id_number=_normalize_query(id_number),
        area=_normalize_query(area),
        uid=_normalize_query(uid),
        month_key=_parse_month_param(month),
        limit=limit,
    )
    return {"attendance": data}


@app.get("/api/export.xlsx")
async def export_xlsx(
    request: Request,
    from_dt: Optional[str] = None,
    to_dt: Optional[str] = None,
    name: Optional[str] = None,
    id_number: Optional[str] = None,
    area: Optional[str] = None,
    uid: Optional[str] = None,
    month: Optional[str] = None,
) -> StreamingResponse:
    _require_admin(request)
    rows = await list_attendance_filtered(
        from_dt=_parse_dt_param(from_dt),
        to_dt=_parse_dt_param(to_dt),
        name=_normalize_query(name),
        id_number=_normalize_query(id_number),
        area=_normalize_query(area),
        uid=_normalize_query(uid),
        month_key=_parse_month_param(month),
        limit=50000,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Control"
    headers = [
        "UID",
        "Nombre",
        "Apellido",
        "Cedula",
        "Telefono",
        "Area",
        "ATR",
        "FechaHora",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row["uid"],
                row["first_name"],
                row["last_name"],
                row["id_number"],
                row["phone"],
                row["area"],
                row["atr"],
                _format_dt(row["read_at"]),
            ]
        )
    ws.freeze_panes = "A2"
    table = Table(displayName="ControlTable", ref=f"A1:H{ws.max_row}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    _autosize_columns(ws)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    headers = {"Content-Disposition": "attachment; filename=control.xlsx"}
    return StreamingResponse(
        stream,
        headers=headers,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/export-people.xlsx")
async def export_people_xlsx(request: Request) -> StreamingResponse:
    _require_admin(request)
    rows = await list_people_all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Personas"
    headers = ["UID", "Nombre", "Apellido", "Cedula", "Telefono", "Area", "Creado"]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row["uid"],
                row["first_name"],
                row["last_name"],
                row["id_number"],
                row["phone"],
                row["area"],
                _format_dt(row["created_at"]),
            ]
        )
    ws.freeze_panes = "A2"
    table = Table(displayName="PersonasTable", ref=f"A1:G{ws.max_row}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    _autosize_columns(ws)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    headers = {"Content-Disposition": "attachment; filename=personas.xlsx"}
    return StreamingResponse(
        stream,
        headers=headers,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/export-people-filtered.xlsx")
async def export_people_filtered_xlsx(
    request: Request,
    from_dt: Optional[str] = None,
    to_dt: Optional[str] = None,
    name: Optional[str] = None,
    id_number: Optional[str] = None,
    area: Optional[str] = None,
) -> StreamingResponse:
    _require_admin(request)
    rows = await list_people_filtered(
        from_dt=_parse_dt_param(from_dt),
        to_dt=_parse_dt_param(to_dt),
        name=_normalize_query(name),
        id_number=_normalize_query(id_number),
        area=_normalize_query(area),
        limit=2000,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Personas"
    headers = ["UID", "Nombre", "Apellido", "Cedula", "Telefono", "Area", "Creado"]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row["uid"],
                row["first_name"],
                row["last_name"],
                row["id_number"],
                row["phone"],
                row["area"],
                _format_dt(row["created_at"]),
            ]
        )
    ws.freeze_panes = "A2"
    table = Table(displayName="PersonasFiltradas", ref=f"A1:G{ws.max_row}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    _autosize_columns(ws)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    headers = {"Content-Disposition": "attachment; filename=personas-filtradas.xlsx"}
    return StreamingResponse(
        stream,
        headers=headers,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
